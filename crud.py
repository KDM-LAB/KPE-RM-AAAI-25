from sqlalchemy.orm import Session
from sqlalchemy import desc, distinct
import models, schemas
from keyphrase_models import model_dict, model_summ_dict
# from similarity_models import similarity_dict
from similarity_models import get_similarity_function_list, similarity_computation, cos_sentbert_vector_representation, cos_glove_cross_vector_representation
from correlations import get_correlations, get_correlations_sync
import numpy as np
import os
from db import engine, sessionLocal
import matplotlib.pyplot as plt
import nltk
from nltk.corpus import stopwords
nltk.download('stopwords')
stop = stopwords.words('english')
import time
import json
from langdetect import detect
import kendalltau_loss
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

def get_reviewers_by_id(db: Session, skip: int, limit: int | None):
    if limit is None:
        return db.query(models.Reviewers).all()
    else:
        return db.query(models.Reviewers).offset(skip).limit(limit).all()

def get_papers_by_id(db: Session, skip: int, limit: int | None):
    if limit is None:
        return db.query(models.Papers).all()
    else:
        return db.query(models.Papers).offset(skip).limit(limit).all()

def extract_papers_keywords(db: Session, model_name: str, skip: int, limit: int | None):
    if limit is None: # extracting keywords from every paper
        results = db.query(models.Papers).all()
    else:
        results = db.query(models.Papers).offset(skip).limit(limit).all()

    promptrank_past_manu_path = r"E:\Backend\reviewer-assignment\promptrank_embeddings\keyphrases\ta"
    json_past_manu_papers = os.listdir(promptrank_past_manu_path)
    promptrank_past_manu_papers = [i.replace(".json", "") for i in json_past_manu_papers]

    for result in results:
        title = result.title
        abstract = result.abstract
        pdf_text_path = result.pdf_text_path

        if model_name == "bertkpe":
            Title = title if isinstance(title, str) else ""
            Abstract = abstract if isinstance(abstract, str) else ""

            if Title or Abstract:
                try:
                    fused_keywords_wo_pdf = model_dict[model_name](title=Title, abstract=Abstract)
                except:
                    print(f"Issue In extraction. title: {Title} || abstract: {Abstract}")
                    fused_keywords_wo_pdf = ";".join([i.lower() for i in Title.split(" ") if i.lower() not in stop])
            else:
                fused_keywords_wo_pdf = ""

            if isinstance(pdf_text_path, str):
                with open(pdf_text_path, "r", encoding="utf-8") as tf:
                    pdf_text = tf.read()
                fused_keywords_w_pdf = model_dict[model_name](pdf_text=pdf_text)
            else:
                fused_keywords_w_pdf = fused_keywords_wo_pdf

        elif model_name == "one2set":
            print(f"paper_pk: {result.paper_pk}", end=" | ")
            Title = title if isinstance(title, str) else ""
            Abstract = abstract if isinstance(abstract, str) else ""

            sta = time.perf_counter()
            if Title or Abstract:
                if Title and Abstract:
                    text = Title + " <eos> " + Abstract
                    fused_keywords_wo_pdf = model_dict[model_name](text=text)
                else:
                    text = Title + Abstract
                    fused_keywords_wo_pdf = model_dict[model_name](text=text)
            else:
                fused_keywords_wo_pdf = ""
            eta = time.perf_counter()
            print(f"titabs: {round(eta-sta, 4)}", end=" | ")

            sp = time.perf_counter()
            if isinstance(pdf_text_path, str):
                with open(pdf_text_path, "r", encoding="utf-8") as tf:
                    pdf_text = tf.read()
                fused_keywords_w_pdf = model_dict[model_name](text=pdf_text)
            else:
                fused_keywords_w_pdf = fused_keywords_wo_pdf
            ep = time.perf_counter()
            print(f"pdf: {round(ep-sp, 4)}", end=" | ")

        # Uncomment this elif block if you wish to populate promptrank keyphrases using precomputed keyphrases
        # elif model_name == "promptrank":
        #     ssid = result.ssId
        #     with open(r'promptrank_keyphrases\promptrank_keywords.json', encoding="utf-8", mode="r") as f:
        #         promptrank_dict = json.load(f)

        #     try:
        #         fused_keywords_wo_pdf = promptrank_dict[ssid]["ta_keywords"]
        #     except:
        #         fused_keywords_wo_pdf = ""

        #     try:
        #         fused_keywords_w_pdf = promptrank_dict[ssid]["pdf_keywords"]
        #         if fused_keywords_w_pdf == "":
        #             fused_keywords_w_pdf = fused_keywords_wo_pdf
        #     except:
        #         fused_keywords_w_pdf = fused_keywords_wo_pdf

        elif model_name in ["keybert_embeddings", "patternrank_embeddings"]:
            ssid = result.ssId
            fused_keywords_wo_pdf = rf"E:\Backend\reviewer-assignment\{model_name}\ta\{ssid}.npy"
            fused_keywords_w_pdf = rf"E:\Backend\reviewer-assignment\{model_name}\pdf\embeddings\{ssid}.npy"

        elif model_name == "promptrank_30":
            ssid = result.ssId
            if ssid is None:
                with open(rf"E:\Backend\reviewer-assignment\promptrank_embeddings\keyphrases\ta\None.json", encoding="utf-8", mode="r") as f:
                    promptrank_dict = json.load(f)
                fused_keywords_wo_pdf = ";".join(promptrank_dict["keyphrases"])
            elif ssid in promptrank_past_manu_papers:
                with open(rf"E:\Backend\reviewer-assignment\promptrank_embeddings\keyphrases\ta\{ssid}.json", encoding="utf-8", mode="r") as f:
                    promptrank_dict = json.load(f)
                fused_keywords_wo_pdf = ";".join(promptrank_dict["keyphrases"])
            else:
                fused_keywords_wo_pdf = None
            # try:
            #     fused_keywords_w_pdf = promptrank_dict[ssid]["pdf_keywords"]
            #     if fused_keywords_w_pdf == "":
            #         fused_keywords_w_pdf = fused_keywords_wo_pdf
            # except:
            #     fused_keywords_w_pdf = fused_keywords_wo_pdf
            fused_keywords_w_pdf = None

        elif model_name == "promptrank_embeddings":
            ssid = result.ssId
            if ssid is None:
                fused_keywords_wo_pdf = rf"E:\Backend\reviewer-assignment\{model_name}\embeddings\ta\None.npy"
            elif ssid in promptrank_past_manu_papers:
                fused_keywords_wo_pdf = rf"E:\Backend\reviewer-assignment\{model_name}\embeddings\ta\{ssid}.npy"
            else:
                fused_keywords_wo_pdf = None
            # try:
            #     fused_keywords_w_pdf = promptrank_dict[ssid]["pdf_keywords"]
            #     if fused_keywords_w_pdf == "":
            #         fused_keywords_w_pdf = fused_keywords_wo_pdf
            # except:
            #     fused_keywords_w_pdf = fused_keywords_wo_pdf
            fused_keywords_w_pdf = None

        elif model_name in model_summ_dict.keys():
            ssid = result.ssId
            mn = model_name.split("_")
            with open(rf'summarization_keywords\{mn[1]}_{mn[0]}_keywords.json', encoding="utf-8", mode="r") as f:
                keywords_dict = json.load(f)

            try:
                fused_keywords_wo_pdf = keywords_dict[ssid]["ta_keywords"]
            except:
                fused_keywords_wo_pdf = ""

            try:
                fused_keywords_w_pdf = keywords_dict[ssid]["pdf_keywords"]
                if fused_keywords_w_pdf == "":
                    fused_keywords_w_pdf = fused_keywords_wo_pdf
            except:
                fused_keywords_w_pdf = fused_keywords_wo_pdf

        else:
            if isinstance(title, str) and isinstance(abstract, str):
                input_string = title + abstract

            elif isinstance(title, str):
                input_string = title

            elif isinstance(abstract, str):
                input_string = abstract

            else:
                input_string = []

            try:
                ta_text_keywords, _, _ = model_dict[model_name](input_string)
            except:
                ta_text_keywords = []

            if isinstance(pdf_text_path, str):
                with open(pdf_text_path, "r", encoding="utf-8") as tf:
                    pdf_text = tf.read()
                for i in range(5): # Trying 5 times to generate keyphrases by truncating input by half each time in case "out of memory" error occurs.
                    try:
                        pdf_text_keywords, _, _ = model_dict[model_name](pdf_text)
                        break
                    except:
                        pdf_text = pdf_text[:len(pdf_text)//2]
                        pdf_text_keywords = []                    
            else:
                pdf_text_keywords = []
            
            fused_keywords_wo_pdf = ";".join(ta_text_keywords)
            if pdf_text_keywords:
                fused_keywords_w_pdf = ";".join(pdf_text_keywords)
            else:
                fused_keywords_w_pdf = fused_keywords_wo_pdf

        kw = models.Model_Paper_Keywords(paper_pk=result.paper_pk, model_name=model_name, model_keywords_wo_pdf=fused_keywords_wo_pdf, model_keywords_w_pdf=fused_keywords_w_pdf)
        db.add(kw)
        db.commit()


def compute_papers_similarity(db: Session, model_name: str, similarity_name: str, skip: int, limit: int | None):
    sim_func_list, sim_args = get_similarity_function_list(similarity_name)
    print(sim_func_list, sim_args)
    if limit is None: # Computing similarity of every record
        results = db.query(models.Rating.reviewer_pk, models.Rating.paper_pk).all()
    else:
        results = db.query(models.Rating.reviewer_pk, models.Rating.paper_pk).offset(skip).limit(limit).all()

    reviewer_cache = results[0].reviewer_pk # setting first reviewer
    past_papers_cache = db.query(models.Model_Paper_Keywords).join(models.Reviewers_Papers, models.Reviewers_Papers.paper_pk == models.Model_Paper_Keywords.paper_pk).filter((models.Reviewers_Papers.reviewer_pk == reviewer_cache) & (models.Model_Paper_Keywords.model_name == model_name)).all()
    for result in results:
        reviewed_paper_data = db.query(models.Model_Paper_Keywords).filter((models.Model_Paper_Keywords.paper_pk == result.paper_pk) & (models.Model_Paper_Keywords.model_name == model_name)).first()
        
        # Note: model_paper_keywords.model_keywords_w_pdf can be empty string if none of title, abstract or pdf_text are available
        reviewer = result.reviewer_pk
        if reviewer == reviewer_cache:
            past_papers_data = past_papers_cache
        else:
            past_papers_data = db.query(models.Model_Paper_Keywords).join(models.Reviewers_Papers, models.Reviewers_Papers.paper_pk == models.Model_Paper_Keywords.paper_pk).filter((models.Reviewers_Papers.reviewer_pk == reviewer) & (models.Model_Paper_Keywords.model_name == model_name)).all()
            past_papers_cache = past_papers_data
            reviewer_cache = reviewer

        author_id = db.query(models.Reviewers).filter(models.Reviewers.reviewer_pk == reviewer).first()
        author_id = author_id.author_id

        with open(rf"E:\gold_standard_github\evaluation_datasets\d_20_1\archives\~{author_id}.jsonl", "r", encoding="utf-8") as f:
            json_data = list(f)
            ta_valid_past_papers = []
            for data in json_data:
                op = json.loads(data)
                try:
                    ta_valid_past_paper_pk = db.query(models.Papers).filter(models.Papers.ssId == op["id"]).first()
                    ta_valid_past_papers.append(ta_valid_past_paper_pk.paper_pk)
                except:
                    continue

        if author_id == "1789029797":
            pdf_valid_past_papers = []
        else:
            with open(rf"E:\gold_standard_github\evaluation_datasets\d_full_20_1\archives\~{author_id}.jsonl", "r", encoding="utf-8") as f:
                json_data = list(f)
                pdf_valid_past_papers = []
                for data in json_data:
                    op = json.loads(data)
                    try:
                        pdf_valid_past_paper_pk = db.query(models.Papers).filter(models.Papers.ssId == op["id"]).first()
                        pdf_valid_past_papers.append(pdf_valid_past_paper_pk.paper_pk)
                    except:
                        continue
        
        if sim_args[7] == "mean":
            if sim_args[0] == "cos-selfembed":
                try:
                    with open(reviewed_paper_data.model_keywords_wo_pdf, "rb") as f:
                        r_em = np.load(f, allow_pickle=True)
                except:
                    r_em = np.array([])
                if r_em.size > 0:
                    similarity_wo_pdf, terms_wo_pdf = 0, 0
                    for past_paper in past_papers_data:
                        if past_paper.paper_pk in ta_valid_past_papers:
                            try:
                                with open(past_paper.model_keywords_wo_pdf, "rb") as f:
                                    p_em = np.load(f, allow_pickle=True)
                            except:
                                continue
                            if p_em.size > 0:
                                similarity_wo_pdf += similarity_computation(r_em, p_em, sim_func_list, sim_args)
                                terms_wo_pdf += 1
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {terms_wo_pdf=}")
                    if terms_wo_pdf == 0:
                        total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_wo_similarity = similarity_wo_pdf/terms_wo_pdf
                else:
                    total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                # print(f"{total_wo_similarity=}")

                try:
                    with open(reviewed_paper_data.model_keywords_w_pdf, "rb") as f:
                        r_em = np.load(f, allow_pickle=True)
                except:
                    r_em = np.array([])
                if r_em.size > 0:
                    similarity_w_pdf, terms_w_pdf = 0, 0
                    for past_paper in past_papers_data:
                        if past_paper.paper_pk in pdf_valid_past_papers:
                            try:
                                with open(past_paper.model_keywords_w_pdf, "rb") as f:
                                    p_em = np.load(f, allow_pickle=True)
                            except:
                                continue
                            if p_em.size > 0:
                                similarity_w_pdf += similarity_computation(r_em, p_em, sim_func_list, sim_args)
                                terms_w_pdf += 1
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {terms_w_pdf=}")
                    if terms_w_pdf == 0:
                        total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_w_similarity = similarity_w_pdf/terms_w_pdf
                else:
                    total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                # print(f"{total_w_similarity=}")

            else: # sim_args[0] in ["jaccard", "cos-glove", "cos-sentbert"]
                if reviewed_paper_data.model_keywords_wo_pdf:
                    similarity_wo_pdf, terms_wo_pdf = 0, 0
                    manuscript_kp = reviewed_paper_data.model_keywords_wo_pdf.split(";")
                    if sim_args[0] == "cos-sentbert":
                        manuscript_kp = [i for i in manuscript_kp if i]
                        vec_m = cos_sentbert_vector_representation(manuscript_kp)
                    if sim_args[0] == "cos-glove":
                        vec_m = cos_glove_cross_vector_representation(manuscript_kp)
                    if sim_args[0] == "jaccard":
                        vec_m = manuscript_kp
                    for past_paper in past_papers_data:
                        if past_paper.model_keywords_wo_pdf and (past_paper.paper_pk in ta_valid_past_papers):
                            simm = similarity_computation(vec_m, past_paper.model_keywords_wo_pdf, sim_func_list, sim_args)
                            if simm is not None:
                                similarity_wo_pdf += simm
                                terms_wo_pdf += 1
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {terms_wo_pdf=}", end=" ")
                    if terms_wo_pdf == 0:
                        total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_wo_similarity = similarity_wo_pdf/terms_wo_pdf
                else:
                    total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                # print(total_wo_similarity)

                if (reviewed_paper_data.model_keywords_w_pdf):
                    similarity_w_pdf, terms_w_pdf = 0, 0
                    manuscript_kp = reviewed_paper_data.model_keywords_w_pdf.split(";")
                    if sim_args[0] == "cos-sentbert":
                        manuscript_kp = [i for i in manuscript_kp if i]
                        vec_m = cos_sentbert_vector_representation(manuscript_kp)
                    if sim_args[0] == "cos-glove":
                        vec_m = cos_glove_cross_vector_representation(manuscript_kp)
                    if sim_args[0] == "jaccard":
                        vec_m = manuscript_kp
                    for past_paper in past_papers_data:
                        if past_paper.model_keywords_w_pdf and (past_paper.paper_pk in pdf_valid_past_papers):
                            simm = similarity_computation(vec_m, past_paper.model_keywords_w_pdf, sim_func_list, sim_args)
                            if simm is not None:
                                similarity_w_pdf += simm
                                terms_w_pdf += 1
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {terms_w_pdf=}", end=" ")
                    if terms_w_pdf == 0:
                        total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_w_similarity = similarity_w_pdf/terms_w_pdf
                else:
                    total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                # print(total_w_similarity, "\n")

        else: # sim_args[7] == "max":
            if sim_args[0] == "cos-selfembed":
                try:
                    with open(reviewed_paper_data.model_keywords_wo_pdf, "rb") as f:
                        r_em = np.load(f, allow_pickle=True)
                except:
                    r_em = np.array([])
                if r_em.size > 0:
                    max_similarity_wo_pdf = -1
                    val_pap_count = 0
                    for past_paper in past_papers_data:
                        if past_paper.paper_pk in ta_valid_past_papers:
                            try:
                                with open(past_paper.model_keywords_wo_pdf, "rb") as f:
                                    p_em = np.load(f, allow_pickle=True)
                            except:
                                continue
                            if p_em.size > 0:
                                sim = similarity_computation(r_em, p_em, sim_func_list, sim_args)
                                val_pap_count += 1
                                if sim > max_similarity_wo_pdf:
                                    max_similarity_wo_pdf = sim
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {val_pap_count=}")
                    if max_similarity_wo_pdf == -1:
                        total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_wo_similarity = max_similarity_wo_pdf
                else:
                    total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
            
                try:
                    with open(reviewed_paper_data.model_keywords_w_pdf, "rb") as f:
                        r_em = np.load(f, allow_pickle=True)
                except:
                    r_em = np.array([])
                if r_em.size > 0:
                    max_similarity_w_pdf = -1
                    val_pap_count = 0
                    for past_paper in past_papers_data:
                        if past_paper.paper_pk in pdf_valid_past_papers:
                            try:
                                with open(past_paper.model_keywords_w_pdf, "rb") as f:
                                    p_em = np.load(f, allow_pickle=True)
                            except:
                                continue
                            if p_em.size > 0:
                                sim = similarity_computation(r_em, p_em, sim_func_list, sim_args)
                                val_pap_count += 1
                                if sim > max_similarity_w_pdf:
                                    max_similarity_w_pdf = sim
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {val_pap_count=}")
                    if max_similarity_w_pdf == -1:
                        total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_w_similarity = max_similarity_w_pdf
                else:
                    total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability

            else: # sim_args[0] in ["jaccard", "cos-glove", "cos-sentbert"]
                if reviewed_paper_data.model_keywords_wo_pdf:
                    max_similarity_wo_pdf = -1
                    val_pap_count = 0
                    manuscript_kp = reviewed_paper_data.model_keywords_wo_pdf.split(";")
                    if sim_args[0] == "cos-sentbert":
                        manuscript_kp = [i for i in manuscript_kp if i]
                        vec_m = cos_sentbert_vector_representation(manuscript_kp)
                    if sim_args[0] == "cos-glove":
                        vec_m = cos_glove_cross_vector_representation(manuscript_kp)
                    if sim_args[0] == "jaccard":
                        vec_m = manuscript_kp
                    for past_paper in past_papers_data:
                        if past_paper.model_keywords_wo_pdf and (past_paper.paper_pk in ta_valid_past_papers):
                            sim = similarity_computation(vec_m, past_paper.model_keywords_wo_pdf, sim_func_list, sim_args)
                            if sim is not None:
                                val_pap_count += 1
                                if sim > max_similarity_wo_pdf:
                                    max_similarity_wo_pdf = sim
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {val_pap_count=}")
                    if max_similarity_wo_pdf == -1:
                        total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_wo_similarity = max_similarity_wo_pdf
                else:
                    total_wo_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability

                if reviewed_paper_data.model_keywords_w_pdf:
                    max_similarity_w_pdf = -1
                    val_pap_count = 0
                    manuscript_kp = reviewed_paper_data.model_keywords_w_pdf.split(";")
                    if sim_args[0] == "cos-sentbert":
                        manuscript_kp = [i for i in manuscript_kp if i]
                        vec_m = cos_sentbert_vector_representation(manuscript_kp)
                    if sim_args[0] == "cos-glove":
                        vec_m = cos_glove_cross_vector_representation(manuscript_kp)
                    if sim_args[0] == "jaccard":
                        vec_m = manuscript_kp
                    for past_paper in past_papers_data:
                        if past_paper.model_keywords_w_pdf and (past_paper.paper_pk in pdf_valid_past_papers):
                            sim = similarity_computation(vec_m, past_paper.model_keywords_w_pdf, sim_func_list, sim_args)
                            if sim is not None:
                                val_pap_count += 1
                                if sim > max_similarity_w_pdf:
                                    max_similarity_w_pdf = sim
                    # print(f"{reviewer_cache=}, {len(past_papers_data)=}, {val_pap_count=}")
                    if max_similarity_w_pdf == -1:
                        total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability
                    else:
                        total_w_similarity = max_similarity_w_pdf
                else:
                    total_w_similarity = None # can't make it zero, cause then it can be infered as the model giving a similarity of 0 instead of data inavailability

        if total_wo_similarity is not None:
            total_wo_similarity = round(total_wo_similarity, 6)
        if total_w_similarity is not None:
            total_w_similarity = round(total_w_similarity, 6)
        similarity = models.Model_Reviewer_Paper_Similarity(reviewer_pk=result.reviewer_pk, paper_pk=result.paper_pk, model_name=model_name, similarity_name=similarity_name, model_similarity_wo_pdf=total_wo_similarity, model_similarity_w_pdf=total_w_similarity)
        db.add(similarity)
        db.commit()

def get_model_extracted_keywords(db: Session, model_name: str, skip: int, limit: int | None):
    available_options = db.query(models.Model_Paper_Keywords.model_name).distinct().all()
    for item in available_options:
        if model_name == item[0]:
            break
    else:
        return {"message":f"available options are {available_options}, choose any from them."}

    if limit is None:
        return db.query(models.Model_Paper_Keywords).filter(models.Model_Paper_Keywords.model_name == model_name).all()
    else:
        return db.query(models.Model_Paper_Keywords).offset(skip).limit(limit).filter(models.Model_Paper_Keywords.model_name == model_name).all()

def get_model_similarity_values(db: Session, model_name: str, similarity_name: str, reviewer_pk: int | None, norm: bool):
    available_options = db.query(models.Model_Reviewer_Paper_Similarity.model_name, models.Model_Reviewer_Paper_Similarity.similarity_name).distinct().all()
    for item in available_options:
        if (model_name == item[0]) and (similarity_name == item[1]):
            break
    else:
        return {"message":f"available options are {available_options}, choose any from them."}

    if reviewer_pk is None:
        results = db.query(models.Rating.reviewer_pk, models.Rating.paper_pk, models.Rating.rating, models.Model_Reviewer_Paper_Similarity.model_similarity_wo_pdf, models.Model_Reviewer_Paper_Similarity.model_similarity_w_pdf) \
            .join(models.Model_Reviewer_Paper_Similarity, (models.Rating.reviewer_pk == models.Model_Reviewer_Paper_Similarity.reviewer_pk) & (models.Rating.paper_pk == models.Model_Reviewer_Paper_Similarity.paper_pk)) \
            .filter((models.Model_Reviewer_Paper_Similarity.model_name == model_name) & (models.Model_Reviewer_Paper_Similarity.similarity_name == similarity_name)) \
            .order_by(desc(models.Model_Reviewer_Paper_Similarity.model_similarity_wo_pdf)).all()
    else:
        results = db.query(models.Rating.reviewer_pk, models.Rating.paper_pk, models.Rating.rating, models.Model_Reviewer_Paper_Similarity.model_similarity_wo_pdf, models.Model_Reviewer_Paper_Similarity.model_similarity_w_pdf) \
            .join(models.Model_Reviewer_Paper_Similarity, (models.Rating.reviewer_pk == models.Model_Reviewer_Paper_Similarity.reviewer_pk) & (models.Rating.paper_pk == models.Model_Reviewer_Paper_Similarity.paper_pk)) \
            .filter((models.Rating.reviewer_pk == reviewer_pk) & (models.Model_Reviewer_Paper_Similarity.model_name == model_name) & (models.Model_Reviewer_Paper_Similarity.similarity_name == similarity_name)) \
            .order_by(desc(models.Model_Reviewer_Paper_Similarity.model_similarity_wo_pdf)).all()

    if norm:
        model_similarity_wo_pdf = []
        model_similarity_w_pdf = []

        model_old_similarity_wo_pdf = [result.model_similarity_wo_pdf for result in results]
        model_old_similarity_w_pdf = [result.model_similarity_w_pdf for result in results]

        minValue_wo_pdf, maxValue_wo_pdf = min(model_old_similarity_wo_pdf), max(model_old_similarity_wo_pdf)
        minValue_w_pdf, maxValue_w_pdf = min(model_old_similarity_w_pdf), max(model_old_similarity_w_pdf)

        for result in results:
            model_similarity_wo_pdf.append(((result.model_similarity_wo_pdf - minValue_wo_pdf)/(maxValue_wo_pdf - minValue_wo_pdf + 1e-2))*(5-0) + 0) # New range [0,5]
            model_similarity_w_pdf.append(((result.model_similarity_w_pdf - minValue_w_pdf)/(maxValue_w_pdf - minValue_w_pdf + 1e-2))*(5-0) + 0)
    else:
        model_similarity_wo_pdf = [result.model_similarity_wo_pdf for result in results]
        model_similarity_w_pdf = [result.model_similarity_w_pdf for result in results]

    # return [{"Reviewer_pk":result.reviewer_pk,
    #         "Paper_pk":result.paper_pk,
    #         "Rating":result.rating,
    #         "Model_Similarity_wo_pdf":round(model_similarity_wo_pdf[idx], 2),
    #         "Model_Similarity_w_pdf":round(model_similarity_w_pdf[idx], 2)}
    #         for idx, result in enumerate(results)]

    return [{"Reviewer_pk":result.reviewer_pk,
            "Paper_pk":result.paper_pk,
            "Rating":result.rating,
            "Model_Similarity_wo_pdf":model_similarity_wo_pdf[idx],
            "Model_Similarity_w_pdf":model_similarity_w_pdf[idx]}
            for idx, result in enumerate(results)]

def null_removal(preds, rates, pids = None):
    p, r, pid = [], [], []
    for idx in range(len(preds)):
        if preds[idx] is not None:
            p.append(preds[idx])
            r.append(rates[idx])
            if pids is not None:
                pid.append(pids[idx])
    if pids is not None:
        return p, r, pid
    return p, r

def get_model_correlation_values(db: Session, model_name: str|None = None, similarity_name: str|None = None, layout: str = "store_correlations", norm: bool = False):
    available_model_similarity_names = db.query(models.Model_Reviewer_Paper_Similarity.model_name, models.Model_Reviewer_Paper_Similarity.similarity_name).distinct().all()
    if (model_name is not None) and (similarity_name is not None):
        for item in available_model_similarity_names:
            if (model_name == item[0]) and (similarity_name == item[1]):
                break
        else:
            return {"message":f"available options are {available_model_similarity_names}, choose any from them."}

    if layout == "whole":
        output = get_model_similarity_values(db=db, model_name=model_name, similarity_name=similarity_name, reviewer_pk=None, norm=norm)
        rating = [row["Rating"] for row in output]
        model_similarity_wo_pdf = [row["Model_Similarity_wo_pdf"] for row in output]
        model_similarity_w_pdf = [row["Model_Similarity_w_pdf"] for row in output]
        model_similarity_wo_pdf, rating_wo = null_removal(model_similarity_wo_pdf, rating)
        model_similarity_w_pdf, rating_w = null_removal(model_similarity_w_pdf, rating)

        return {"Model_Correlation_wo_pdf": get_correlations(rating_wo, model_similarity_wo_pdf),
                "Model_Correlation_w_pdf": get_correlations(rating_w, model_similarity_w_pdf)}

    elif layout == "by_reviewer":
        return_result = {}
        length = {}
        rating_result = {}
        for idx in range(1,59,1):
            output = get_model_similarity_values(db=db, model_name=model_name, similarity_name=similarity_name, reviewer_pk=idx, norm=norm)
            rating = [row["Rating"] for row in output]
            paper_pks = [row["Paper_pk"] for row in output]
            model_similarity_wo_pdf = [row["Model_Similarity_wo_pdf"] for row in output]
            # model_similarity_w_pdf = [row["Model_Similarity_w_pdf"] for row in output]
            model_similarity_wo_pdf, rating_wo, paper_pks_wo = null_removal(model_similarity_wo_pdf, rating, paper_pks)
            # model_similarity_w_pdf, rating_w = null_removal(model_similarity_w_pdf, rating)

            # combined_result = {"Model_Correlation_wo_pdf": get_correlations(rating_wo, model_similarity_wo_pdf),
            #                 "Model_Correlation_w_pdf": get_correlations(rating_w, model_similarity_w_pdf)}
            # return_result[idx] = combined_result
            rating_result[idx] = dict(zip(paper_pks_wo, zip(rating_wo, model_similarity_wo_pdf)))
            # length[idx] = len(rating)
        # print({k:(return_result[k]["Model_Correlation_wo_pdf"]["Pearson"]["Correlation"], length[k]) for k in return_result.keys()})
        print(rating_result)
        return return_result

    elif layout == "by_reviewer_sync":
        wo_pdf_corr = {"pearson":[], "spearman":[], "kendaltau":[]}
        w_pdf_corr = {"pearson":[], "spearman":[], "kendaltau":[]}
        for idx in range(1,59,1):
            output = get_model_similarity_values(db=db, model_name=model_name, similarity_name=similarity_name, reviewer_pk=idx, norm=norm)
            rating = [row["Rating"] for row in output]
            model_similarity_wo_pdf = [row["Model_Similarity_wo_pdf"] for row in output]
            model_similarity_w_pdf = [row["Model_Similarity_w_pdf"] for row in output]

            p, s, k = get_correlations_sync(rating, model_similarity_wo_pdf)
            wo_pdf_corr["pearson"].append(p)
            wo_pdf_corr["spearman"].append(s)
            wo_pdf_corr["kendaltau"].append(k)

            p, s, k = get_correlations_sync(rating, model_similarity_w_pdf)
            w_pdf_corr["pearson"].append(p)
            w_pdf_corr["spearman"].append(s)
            w_pdf_corr["kendaltau"].append(k)

        for corr in [wo_pdf_corr, w_pdf_corr]:
            for k in corr.keys():
                corr[k] = np.mean(corr[k])

        return {"Model_Correlation_wo_pdf": wo_pdf_corr,
                "Model_Correlation_w_pdf": w_pdf_corr}
    
    elif layout == "by_reviewer_describe":
        wo_pdf_dict = {"Pearson":[], "Spearman":[], "Kendalltau":[]}
        w_pdf_dict = {"Pearson":[], "Spearman":[], "Kendalltau":[]}

        refs_wo = {} # for kendalltau loss
        refs_w = {} # for kendalltau loss
        preds_wo_pdf = {} # for kendalltau loss
        preds_w_pdf = {} # for kendalltau loss

        for idx in range(1,59,1):
            output = get_model_similarity_values(db=db, model_name=model_name, similarity_name=similarity_name, reviewer_pk=idx, norm=norm)

            rating = [row["Rating"] for row in output]
            model_similarity_wo_pdf = [row["Model_Similarity_wo_pdf"] for row in output]
            model_similarity_w_pdf = [row["Model_Similarity_w_pdf"] for row in output]
            paper_id = [str(row["Paper_pk"]) for row in output]
            model_similarity_wo_pdf, rating_wo, paper_id_wo = null_removal(model_similarity_wo_pdf, rating, paper_id)
            model_similarity_w_pdf, rating_w, paper_id_w = null_removal(model_similarity_w_pdf, rating, paper_id)

            refs_wo[idx] = dict(zip(paper_id_wo, rating_wo))
            refs_w[idx] = dict(zip(paper_id_w, rating_w))
            preds_wo_pdf[idx] = dict(zip(paper_id_wo, model_similarity_wo_pdf))
            preds_w_pdf[idx] = dict(zip(paper_id_w, model_similarity_w_pdf))

            if (wo_pdf_data := get_correlations(rating_wo, model_similarity_wo_pdf)) is not None:
                wo_pdf_dict["Pearson"].append(wo_pdf_data["Pearson"]["Correlation"])
                wo_pdf_dict["Spearman"].append(wo_pdf_data["Spearman"]["Correlation"])
                wo_pdf_dict["Kendalltau"].append(wo_pdf_data["Kendalltau"]["Correlation"])

            if (w_pdf_data := get_correlations(rating_w, model_similarity_w_pdf)) is not None:
                w_pdf_dict["Pearson"].append(w_pdf_data["Pearson"]["Correlation"])
                w_pdf_dict["Spearman"].append(w_pdf_data["Spearman"]["Correlation"])
                w_pdf_dict["Kendalltau"].append(w_pdf_data["Kendalltau"]["Correlation"])

        # Uncomment to plot and save a graph for pearson correlation of different reviewers
        # plt.figure()
        # plt.scatter(list(range(1,59,1)), wo_pdf_dict["Pearson"])
        # title = f"{model_name} model with {similarity_name} similarity"
        # plt.title(title)
        # plt.xlabel("Reviewer ID")
        # plt.ylabel("Pearson Correlation")
        # plt.savefig(f".\correlation_graphs\{title}.jpg", dpi=400)

        ktl_wo_pdf = kendalltau_loss.compute_main_metric(preds_wo_pdf, refs_wo)
        ktl_w_pdf = kendalltau_loss.compute_main_metric(preds_w_pdf, refs_w)

        # return {"wo_pdf":{"kendalTauLoss":ktl_wo_pdf,
        #                 "mean":{"Pearson":np.mean(wo_pdf_dict["Pearson"]), "Spearman":np.mean(wo_pdf_dict["Spearman"]), "Kendalltau":np.mean(wo_pdf_dict["Kendalltau"])},
        #                 "median":{"Pearson":np.median(wo_pdf_dict["Pearson"]), "Spearman":np.median(wo_pdf_dict["Spearman"]), "Kendalltau":np.median(wo_pdf_dict["Kendalltau"])},
        #                 "std":{"Pearson":np.std(wo_pdf_dict["Pearson"]), "Spearman":np.std(wo_pdf_dict["Spearman"]), "Kendalltau":np.std(wo_pdf_dict["Kendalltau"])}},
        #         "w_pdf":{"kendalTauLoss":ktl_w_pdf,
        #                 "mean":{"Pearson":np.mean(w_pdf_dict["Pearson"]), "Spearman":np.mean(w_pdf_dict["Spearman"]), "Kendalltau":np.mean(w_pdf_dict["Kendalltau"])},
        #                 "median":{"Pearson":np.median(w_pdf_dict["Pearson"]), "Spearman":np.median(w_pdf_dict["Spearman"]), "Kendalltau":np.median(w_pdf_dict["Kendalltau"])},
        #                 "std":{"Pearson":np.std(w_pdf_dict["Pearson"]), "Spearman":np.std(w_pdf_dict["Spearman"]), "Kendalltau":np.std(w_pdf_dict["Kendalltau"])}}}
        print({"wo_pdf":{"kendalTauLoss":ktl_wo_pdf,
                "mean":{"Pearson":np.mean(wo_pdf_dict["Pearson"]), "Spearman":np.mean(wo_pdf_dict["Spearman"]), "Kendalltau":np.mean(wo_pdf_dict["Kendalltau"])},
                "median":{"Pearson":np.median(wo_pdf_dict["Pearson"]), "Spearman":np.median(wo_pdf_dict["Spearman"]), "Kendalltau":np.median(wo_pdf_dict["Kendalltau"])},
                "std":{"Pearson":np.std(wo_pdf_dict["Pearson"]), "Spearman":np.std(wo_pdf_dict["Spearman"]), "Kendalltau":np.std(wo_pdf_dict["Kendalltau"])}},
        "w_pdf":{"kendalTauLoss":ktl_w_pdf,
                "mean":{"Pearson":np.mean(w_pdf_dict["Pearson"]), "Spearman":np.mean(w_pdf_dict["Spearman"]), "Kendalltau":np.mean(w_pdf_dict["Kendalltau"])},
                "median":{"Pearson":np.median(w_pdf_dict["Pearson"]), "Spearman":np.median(w_pdf_dict["Spearman"]), "Kendalltau":np.median(w_pdf_dict["Kendalltau"])},
                "std":{"Pearson":np.std(w_pdf_dict["Pearson"]), "Spearman":np.std(w_pdf_dict["Spearman"]), "Kendalltau":np.std(w_pdf_dict["Kendalltau"])}}})

    elif layout == "store_correlations":
        if os.path.exists("correlation_results.xlsx"):
            wb = load_workbook("correlation_results.xlsx")
            ws = wb['model_similarity_name']
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "model_similarity_name"
            ws.append(["model_name", "similarity_name"])
            ws['A1'].font = Font(bold=True)
            ws['B1'].font = Font(bold=True)

        present_model_similarity_names = set()
        for idx, row in enumerate(ws):
            if idx > 0:
                present_model_similarity_names.add((row[0].value, row[1].value))
        new_model_similarity_set = set(available_model_similarity_names) - present_model_similarity_names

        print(f"{len(available_model_similarity_names)=}")
        print(f"{len(present_model_similarity_names)=}")
        print(f"{len(new_model_similarity_set)=}")

        for pair in new_model_similarity_set:
            print(f"Computing for {pair}")
            sim_args = pair[1].split("_")
            ws.append([pair[0], pair[1]])
            try:
                mws = wb[pair[0]] # model worksheet
            except:
                wb.create_sheet(pair[0])
                mws = wb[pair[0]]
                mws_title_list = ["sim_method", "type", "discard", "lemma", "#kp", "#past papers", "pos weight factor", "mode", "format", "pearson", "spearman", "kendal tau", "kendal tau loss"]
                mws.append(mws_title_list)
                for i in range(1, len(mws_title_list)+1):
                    mws[get_column_letter(i) + '1'].font = Font(bold=True)

            wo_pdf_dict = {"Pearson":[], "Spearman":[], "Kendalltau":[]}
            w_pdf_dict = {"Pearson":[], "Spearman":[], "Kendalltau":[]}

            refs_wo = {} # for kendalltau loss
            refs_w = {} # for kendalltau loss
            preds_wo_pdf = {} # for kendalltau loss
            preds_w_pdf = {} # for kendalltau loss

            for idx in range(1,59,1):
                output = get_model_similarity_values(db=db, model_name=pair[0], similarity_name=pair[1], reviewer_pk=idx, norm=norm)

                rating = [row["Rating"] for row in output]
                model_similarity_wo_pdf = [row["Model_Similarity_wo_pdf"] for row in output]
                model_similarity_w_pdf = [row["Model_Similarity_w_pdf"] for row in output]
                paper_id = [str(row["Paper_pk"]) for row in output]
                model_similarity_wo_pdf, rating_wo, paper_id_wo = null_removal(model_similarity_wo_pdf, rating, paper_id)
                model_similarity_w_pdf, rating_w, paper_id_w = null_removal(model_similarity_w_pdf, rating, paper_id)

                refs_wo[idx] = dict(zip(paper_id_wo, rating_wo))
                refs_w[idx] = dict(zip(paper_id_w, rating_w))
                preds_wo_pdf[idx] = dict(zip(paper_id_wo, model_similarity_wo_pdf))
                preds_w_pdf[idx] = dict(zip(paper_id_w, model_similarity_w_pdf))

                if (wo_pdf_data := get_correlations(rating_wo, model_similarity_wo_pdf)) is not None:
                    wo_pdf_dict["Pearson"].append(wo_pdf_data["Pearson"]["Correlation"])
                    wo_pdf_dict["Spearman"].append(wo_pdf_data["Spearman"]["Correlation"])
                    wo_pdf_dict["Kendalltau"].append(wo_pdf_data["Kendalltau"]["Correlation"])
                else:
                    print(f"flaw in reviewer {idx} in {pair[0]} model and {pair[1]} similarity | title+abs format")

                if (w_pdf_data := get_correlations(rating_w, model_similarity_w_pdf)) is not None:
                    w_pdf_dict["Pearson"].append(w_pdf_data["Pearson"]["Correlation"])
                    w_pdf_dict["Spearman"].append(w_pdf_data["Spearman"]["Correlation"])
                    w_pdf_dict["Kendalltau"].append(w_pdf_data["Kendalltau"]["Correlation"])
                else:
                    print(f"flaw in reviewer {idx} in {pair[0]} model and {pair[1]} similarity | pdf format")

            ktl_wo_pdf = kendalltau_loss.compute_main_metric(preds_wo_pdf, refs_wo)
            ktl_w_pdf = kendalltau_loss.compute_main_metric(preds_w_pdf, refs_w)

            corr_wo_list = []
            for i in ["Pearson", "Spearman", "Kendalltau"]:
                corr_wo_list.append(round(np.mean(wo_pdf_dict[i]), 4))

            corr_w_list = []
            for i in ["Pearson", "Spearman", "Kendalltau"]:
                corr_w_list.append(round(np.mean(w_pdf_dict[i]), 4))

            skip_row = [None for i in range(mws.max_column)]

            mws.append(sim_args + ["title+abs"] + corr_wo_list + [round(ktl_wo_pdf, 4)])
            mws.append(sim_args + ["pdf text"] + corr_w_list + [round(ktl_w_pdf, 4)])
            mws.append(skip_row)

        wb.save("correlation_results.xlsx")

    else:
        raise Exception("Unknown format, enter either 'whole', 'by_reviewer_sync', 'by_reviewer_describe' or 'by_reviewer'")

# starting local session to directly run crud.py (do not comment)
db = sessionLocal()

# ------------------------------------------------------------------------------------- #
# SECTION 0
# Uncomment SECTION 0 and comment all other SECTIONS to test similarity computation code
# compute_papers_similarity(db, f"positionrank", f"cos-glove_cross_nd_nl_5_20_0_mean", skip=0, limit=None)
# BE SURE TO DELETE THE RECORDS GENERATED IN THE DB CORRESPONDING TO THE ABOVE LINE BEFORE PROCEDING FURTHER
# ------------------------------------------------------------------------------------- #


# ------------------------------------------------------------------------------------- #
# SECTION 1
# Uncomment SECTION 1 and comment all other SECTIONS to compute and store similarity scores in db
# cos-glove calculation for the below models
# for m in ["positionrank", "multipartiterank", "yake"]:
#     for s in ["cos-glove"]:
#         for k in [30, 15, 5]:
#             for mod in ["mean", "max"]:
#                 compute_papers_similarity(db, f"{m}", f"{s}_cross_nd_nl_{k}_20_0_{mod}", skip=0, limit=None)
# for m in ["positionrank", "multipartiterank", "yake"]:
#     for s in ["cos-glove"]:
#         for k in [30, 15, 5, -30, -15, -5]:
#             for mod in ["mean", "max"]:
#                 compute_papers_similarity(db, f"{m}", f"{s}_cross_d_nl_{k}_20_0_{mod}", skip=0, limit=None)

# # cos-sentbert calculation for below models
# for m in ["textrank", "singlerank", "positionrank", "multipartiterank", "yake"]:
#     for s in ["cos-sentbert"]:
#         for k in [30, 15, 5]:
#             for mod in ["mean", "max"]:
#                 compute_papers_similarity(db, f"{m}", f"{s}_cross_nd_nl_{k}_20_0_{mod}", skip=0, limit=None)
# for m in ["textrank", "singlerank", "positionrank", "multipartiterank", "yake"]:
#     for s in ["cos-sentbert"]:
#         for k in [30, 15, 5, -30, -15, -5]:
#             for mod in ["mean", "max"]:
#                 compute_papers_similarity(db, f"{m}", f"{s}_cross_d_nl_{k}_20_0_{mod}", skip=0, limit=None)
# ------------------------------------------------------------------------------------- #


# ------------------------------------------------------------------------------------- #
# SECTION 2
# Uncomment SECTION 2 and comment all other SECTIONS to compute all 4 scores and fill the excel sheet automatically
# get_model_correlation_values(db)
# ------------------------------------------------------------------------------------- #


# ------------------------------------------------------------------------------------- #
# SECTION 3
# Uncomment SECTION 3 and comment all other SECTIONS to compute keyphrases for promptrank
# extract_papers_keywords(db, model_name="promptrank", skip=0, limit=None)
# ------------------------------------------------------------------------------------- #


# ------------------------------------------------------------------------------------- #
# SECTION 4
# Uncomment SECTION 4 and comment all other SECTIONS to compute and store similarity scores in db for promptrank
# for m in ["promptrank"]:
#     for s in ["jaccard", "cos-glove", "cos-sentbert"]:
#         for k in [30, 15, 5]:
#             for mod in ["mean", "max"]:
#                 compute_papers_similarity(db, f"{m}", f"{s}_cross_nd_nl_{k}_20_0_{mod}", skip=0, limit=None)
# for m in ["promptrank"]:
#     for s in ["jaccard", "cos-glove", "cos-sentbert"]:
#         for k in [30, 15, 5, -30, -15, -5]:
#             for mod in ["mean", "max"]:
#                 compute_papers_similarity(db, f"{m}", f"{s}_cross_d_nl_{k}_20_0_{mod}", skip=0, limit=None)
# ------------------------------------------------------------------------------------- #


# ------------------------------------------------------------------------------------- #
# SECTION 5
# Uncomment SECTION 5 and comment all other SECTIONS to compute all 4 scores and fill the excel sheet automatically (for promptrank)
# get_model_correlation_values(db)
# ------------------------------------------------------------------------------------- #

db.close()